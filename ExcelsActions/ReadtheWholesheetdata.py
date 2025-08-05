import openpyxl

path="C:\\Users\\Lenovo\\PycharmProjects\\Automation-office\\Excels\\Testingdata.xlsx"

workbook=openpyxl.load_workbook(path)

sheet=workbook["DemoData"]

row=sheet.max_row
column=sheet.max_column
print(row) #it gives total row
print(column) #it gives total columns

#Approch 1
for i in range(1,row+1):
    for j in range(1,column+1):
        print(sheet.cell(row=i, column=j).value)

#Approach 2: Read a fixed range like A1:B4 using sheet["A1:B4"]
# This returns a tuple of tuples (rows of cells)
for i in sheet['a1':'b4']:
    for j in i:
        print(j.value)