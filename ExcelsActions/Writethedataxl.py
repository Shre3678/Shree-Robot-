import openpyxl

# workbook =openpyxl.Workbook()
# sheet=workbook.active
# sheet.title = "Demo Dat"   #setting the name of sheet
# workbook.save("DemoData.xlsx") #it will crate the xl
# print(sheet.title)

#write some information on particular cell
#workbook1 =openpyxl.Workbook()
#sheet1=workbook1.active
# sheet1.title = "Demo Dat"   #setting the name of sheet
# workbook1.save("d:\\DemoData.xlsx") #it will crate the xl
# print(sheet1.title)
# sheet1['A2'].value = "Demo Information"
# workbook1.save("d:\\DemoData.xlsx")

#Create the sheet and write
#workbook1 =openpyxl.Workbook()
#worksheet1 = workbook1.create_sheet(title="Testing")
# worksheet1['A3'].value="538953"
# workbook1.save("d:\\DemoData.xlsx")

#remove the sheet
workbook1 =openpyxl.Workbook()
worksheet1 = workbook1.create_sheet(title="Testing1")
workbook1.remove(worksheet1["Testing1"])
workbook1.save("D:\\DemoData.xlsx")
