import openpyxl


workbook=openpyxl.load_workbook("C:\\Users\\Lenovo\\PycharmProjects\\Automation-office\\Excels\\Testingdata.xlsx")

#Approch 1
sh=workbook["DemoData"]
print(sh['a2'].value)

#Approch 2
c1=sh.cell(3,2)
print(c1.value)
print(c1.row)  # To know which row the value is coming from
print(c1.column) # To know which column the value is coming from

#with keyword mentions
c2=sh.cell(column=2, row=3)
print(c2.value)



