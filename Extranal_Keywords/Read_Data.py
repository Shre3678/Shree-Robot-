import openpyxl

workbook=openpyxl.load_workbook("C:\\Users\\Lenovo\\PycharmProjects\\Automation-office\\Excels\\Testingdata.xlsx")

def fetch_countof_rows(sheetname):
    sheet=workbook["DemoData"]
    return  sheet.max_row

def fetch_cell_data(Sheetname, row, cell):
    sheet=workbook["DemoData"]
    cell=sheet.cell(int(row), int(cell))
    return cell.value

#print(fetch_countof_rows("DemoData"))
#print(fetch_cell_data(2, 2))