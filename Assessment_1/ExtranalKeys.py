import openpyxl

#from Extranal_Keywords.Read_Data import workbook

workbook=openpyxl.load_workbook("C:\\Users\\Lenovo\\PycharmProjects\\Automation-office\\Excels\\Pepperfry.xlsx")

def fetch_total_rows(sheetname):
    sheet=workbook["Items"]
    return  sheet.max_row

def fetch_cell_datas(Sheetname, row, cell):
    sheet=workbook["Items"]
    cell=sheet.cell(int(row), int(cell))
    return cell.value